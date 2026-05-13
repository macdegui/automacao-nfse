"""
Cria o arquivo config.xlsx com a estrutura de mapeamento de empresas.
Execute este script uma vez para gerar o arquivo de configuração.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def criar_config():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapeamento"

    # Estilo cabeçalho
    fill_header = PatternFill("solid", fgColor="1A6FD4")
    font_header = Font(bold=True, color="FFFFFF", size=12)
    borda = Border(
        left=Side(style='thin', color='1E3A5F'),
        right=Side(style='thin', color='1E3A5F'),
        top=Side(style='thin', color='1E3A5F'),
        bottom=Side(style='thin', color='1E3A5F')
    )

    # Cabeçalhos
    headers = ["Nome no Portal NFSe", "Nome na Pasta do Servidor"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = borda

    # Exemplos
    exemplos = [
        ("EL COMMERCE UTILIDADES E SERVICOS LTDA", "El Commerce"),
        ("EMPRESA EXEMPLO LTDA", "Empresa Exemplo"),
    ]

    fill_exemplo = PatternFill("solid", fgColor="F0F4FF")
    for row, (portal, servidor) in enumerate(exemplos, 2):
        for col, valor in enumerate([portal, servidor], 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.alignment = Alignment(vertical='center')
            cell.border = borda
            if row % 2 == 0:
                cell.fill = fill_exemplo

    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 35
    ws.row_dimensions[1].height = 28

    # Instruções na aba 2
    ws2 = wb.create_sheet("Instruções")
    instrucoes = [
        ("INSTRUÇÕES DE USO", True),
        ("", False),
        ("1. Na coluna 'Nome no Portal NFSe':", False),
        ("   Digite o nome EXATAMENTE como aparece na pasta do Download.", False),
        ("   Ex: EL COMMERCE UTILIDADES E SERVICOS LTDA", False),
        ("", False),
        ("2. Na coluna 'Nome na Pasta do Servidor':", False),
        ("   Digite o nome da pasta correspondente no servidor Z:", False),
        ("   Ex: El Commerce", False),
        ("", False),
        ("3. A comparação ignora maiúsculas/minúsculas.", False),
        ("4. Adicione uma linha por empresa.", False),
        ("5. Salve o arquivo após cada alteração.", False),
    ]

    for row, (texto, negrito) in enumerate(instrucoes, 1):
        cell = ws2.cell(row=row, column=1, value=texto)
        if negrito:
            cell.font = Font(bold=True, size=13, color="1A6FD4")
        else:
            cell.font = Font(size=11)

    ws2.column_dimensions['A'].width = 60

    # Salva
    caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.xlsx")
    wb.save(caminho)
    print(f"✅ Arquivo criado: {caminho}")
    return caminho


if __name__ == "__main__":
    criar_config()
